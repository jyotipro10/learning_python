import openpyxl 

wb = openpyxl.Workbook() 
  

sheet = wb.active 
  

c1 = sheet.cell(row = 1, column = 1) 
  

c1.value = "First Name"
  
c2 = sheet.cell(row= 1 , column = 2) 
c2.value = "Last Name"

c3 = sheet['C1'] 
c3.value = "Roll No."
  

c4 = sheet['D1'] 
c4.value = "Department"

c5 = sheet['A2']
c5.value = "Jyotiprakash"

c6 = sheet['B2']
c6.value = "Das Karmakar"

c7 = sheet['C2']
c7.value = "180333"

c8 = sheet['D2']
c8.value = "Civil Engineering"
  

wb.save("C:\\Users\\HP\\Desktop\\xlp.xlsx") 
